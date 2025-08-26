"""
Core utilities for EduCore Ultra project.
"""
import logging
import hashlib
import uuid
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Union
from functools import wraps
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.http import HttpRequest, HttpResponse
from django.utils import timezone
from django.conf import settings
from rest_framework import status
from rest_framework.response import Response

logger = logging.getLogger(__name__)


def generate_unique_id(prefix: str = "") -> str:
    """Generate a unique identifier."""
    unique_id = str(uuid.uuid4()).replace('-', '')[:12]
    return f"{prefix}{unique_id}" if prefix else unique_id


def hash_string(text: str, algorithm: str = 'sha256') -> str:
    """Hash a string using the specified algorithm."""
    hash_func = getattr(hashlib, algorithm, hashlib.sha256)
    return hash_func(text.encode()).hexdigest()


def format_file_size(size_bytes: int) -> str:
    """Format file size in human readable format."""
    if size_bytes == 0:
        return "0B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f}{size_names[i]}"


def validate_email(email: str) -> bool:
    """Validate email format."""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_phone(phone: str) -> bool:
    """Validate phone number format."""
    import re
    # Remove all non-digit characters
    digits_only = re.sub(r'\D', '', phone)
    # Check if it's between 10-15 digits
    return 10 <= len(digits_only) <= 15


def cache_result(timeout: int = 300):
    """Decorator to cache function results."""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Create cache key from function name and arguments
            cache_key = f"{func.__name__}:{hash(str(args) + str(sorted(kwargs.items())))}"
            
            # Try to get from cache
            result = cache.get(cache_key)
            if result is not None:
                return result
            
            # Execute function and cache result
            result = func(*args, **kwargs)
            cache.set(cache_key, result, timeout)
            return result
        return wrapper
    return decorator


def rate_limit(max_requests: int = 100, window: int = 3600):
    """Decorator to implement rate limiting."""
    def decorator(func):
        @wraps(func)
        def wrapper(request: HttpRequest, *args, **kwargs):
            # Get client IP
            client_ip = get_client_ip(request)
            cache_key = f"rate_limit:{client_ip}:{func.__name__}"
            
            # Get current request count
            request_count = cache.get(cache_key, 0)
            
            if request_count >= max_requests:
                return Response(
                    {"error": "Rate limit exceeded"},
                    status=status.HTTP_429_TOO_MANY_REQUESTS
                )
            
            # Increment request count
            cache.set(cache_key, request_count + 1, window)
            
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


def get_client_ip(request: HttpRequest) -> str:
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_activity(user_id: int, action: str, details: Dict[str, Any] = None):
    """Log user activity."""
    try:
        from apps.analytics.models import UserActivity
        UserActivity.objects.create(
            user_id=user_id,
            action=action,
            details=details or {}
        )
    except Exception as e:
        logger.error(f"Failed to log activity: {e}")


def activity_logger(action: str):
    """Decorator to log user activity."""
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            result = func(request, *args, **kwargs)
            
            # Log activity if user is authenticated
            if hasattr(request, 'user') and request.user.is_authenticated:
                log_activity(
                    user_id=request.user.id,
                    action=action,
                    details={
                        'method': request.method,
                        'path': request.path,
                        'status_code': getattr(result, 'status_code', None)
                    }
                )
            
            return result
        return wrapper
    return decorator


def paginate_queryset(queryset, page: int = 1, page_size: int = 20):
    """Paginate a queryset."""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(queryset, page_size)
    
    try:
        items = paginator.page(page)
    except PageNotAnInteger:
        items = paginator.page(1)
    except EmptyPage:
        items = paginator.page(paginator.num_pages)
    
    return {
        'items': items,
        'total_pages': paginator.num_pages,
        'total_count': paginator.count,
        'has_next': items.has_next(),
        'has_previous': items.has_previous(),
        'current_page': items.number
    }


def export_to_csv(queryset, fields: List[str], filename: str = None):
    """Export queryset to CSV format."""
    import csv
    from io import StringIO
    from django.http import HttpResponse
    
    if not filename:
        filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(fields)
    
    # Write data
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field, '')
            if hasattr(value, '__call__'):
                value = value()
            row.append(str(value))
        writer.writerow(row)
    
    return response


def export_to_excel(queryset, fields: List[str], filename: str = None):
    """Export queryset to Excel format."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Write header
        for col, field in enumerate(fields, 1):
            ws.cell(row=1, column=col, value=field)
        
        # Write data
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = getattr(obj, field, '')
                if hasattr(value, '__call__'):
                    value = value()
                ws.cell(row=row, column=col, value=str(value))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        raise ImportError("openpyxl is required for Excel export")


def send_email_notification(
    to_email: str,
    subject: str,
    template_name: str,
    context: Dict[str, Any] = None,
    from_email: str = None
):
    """Send email notification using Django's email system."""
    try:
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        
        if not from_email:
            from_email = settings.DEFAULT_FROM_EMAIL
        
        # Render email template
        html_message = render_to_string(f'emails/{template_name}.html', context or {})
        plain_message = render_to_string(f'emails/{template_name}.txt', context or {})
        
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=from_email,
            recipient_list=[to_email],
            html_message=html_message,
            fail_silently=False
        )
        
        logger.info(f"Email sent to {to_email}: {subject}")
        
    except Exception as e:
        logger.error(f"Failed to send email to {to_email}: {e}")


def generate_password_reset_token(user_id: int) -> str:
    """Generate a password reset token."""
    import secrets
    token = secrets.token_urlsafe(32)
    cache_key = f"password_reset:{token}"
    cache.set(cache_key, user_id, timeout=3600)  # 1 hour expiry
    return token


def verify_password_reset_token(token: str) -> Optional[int]:
    """Verify a password reset token and return user ID."""
    cache_key = f"password_reset:{token}"
    user_id = cache.get(cache_key)
    if user_id:
        cache.delete(cache_key)  # Use once
        return user_id
    return None


def sanitize_filename(filename: str) -> str:
    """Sanitize filename for safe storage."""
    import re
    # Remove or replace unsafe characters
    filename = re.sub(r'[^\w\-_\.]', '_', filename)
    # Limit length
    if len(filename) > 255:
        name, ext = filename.rsplit('.', 1)
        filename = name[:255-len(ext)-1] + '.' + ext
    return filename


def validate_file_upload(file, allowed_types: List[str] = None, max_size: int = None):
    """Validate file upload."""
    if allowed_types is None:
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'application/pdf']
    
    if max_size is None:
        max_size = 5 * 1024 * 1024  # 5MB
    
    # Check file type
    if file.content_type not in allowed_types:
        raise ValidationError(f"File type {file.content_type} is not allowed")
    
    # Check file size
    if file.size > max_size:
        raise ValidationError(f"File size {file.size} exceeds maximum allowed size {max_size}")
    
    return True


def get_file_extension(filename: str) -> str:
    """Get file extension from filename."""
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def is_valid_date(date_string: str, format: str = '%Y-%m-%d') -> bool:
    """Check if a date string is valid."""
    try:
        datetime.strptime(date_string, format)
        return True
    except ValueError:
        return False


def get_date_range(start_date: str, end_date: str, format: str = '%Y-%m-%d') -> List[str]:
    """Get list of dates between start and end date."""
    start = datetime.strptime(start_date, format)
    end = datetime.strptime(end_date, format)
    
    date_list = []
    current = start
    while current <= end:
        date_list.append(current.strftime(format))
        current += timedelta(days=1)
    
    return date_list


def format_currency(amount: float, currency: str = 'USD') -> str:
    """Format currency amount."""
    import locale
    try:
        locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
        return locale.currency(amount, grouping=True, symbol=currency)
    except:
        return f"{currency} {amount:,.2f}"


def mask_sensitive_data(data: str, mask_char: str = '*') -> str:
    """Mask sensitive data like credit card numbers or phone numbers."""
    if len(data) <= 4:
        return data
    
    return data[:2] + mask_char * (len(data) - 4) + data[-2:]


def generate_otp(length: int = 6) -> str:
    """Generate a numeric OTP."""
    import random
    return ''.join([str(random.randint(0, 9)) for _ in range(length)])


def validate_otp(otp: str, stored_otp: str) -> bool:
    """Validate OTP."""
    return otp == stored_otp


def get_system_info() -> Dict[str, Any]:
    """Get system information."""
    import platform
    import psutil
    
    return {
        'platform': platform.platform(),
        'python_version': platform.python_version(),
        'cpu_count': psutil.cpu_count(),
        'memory_total': psutil.virtual_memory().total,
        'memory_available': psutil.virtual_memory().available,
        'disk_usage': psutil.disk_usage('/').percent
    }


def cleanup_old_files(directory: str, days_old: int = 30):
    """Clean up old files from a directory."""
    import os
    from pathlib import Path
    
    cutoff_date = timezone.now() - timedelta(days=days_old)
    directory_path = Path(directory)
    
    if not directory_path.exists():
        return
    
    for file_path in directory_path.rglob('*'):
        if file_path.is_file():
            file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
            if file_mtime < cutoff_date:
                try:
                    file_path.unlink()
                    logger.info(f"Deleted old file: {file_path}")
                except Exception as e:
                    logger.error(f"Failed to delete file {file_path}: {e}")


def backup_database():
    """Create database backup."""
    import subprocess
    import os
    
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(backup_dir, f'backup_{timestamp}.sql')
    
    try:
        if settings.DATABASES['default']['ENGINE'] == 'django.db.backends.postgresql':
            cmd = [
                'pg_dump',
                '-h', settings.DATABASES['default']['HOST'],
                '-U', settings.DATABASES['default']['USER'],
                '-d', settings.DATABASES['default']['NAME'],
                '-f', backup_file
            ]
            subprocess.run(cmd, check=True)
        else:
            # SQLite backup
            import shutil
            db_file = settings.DATABASES['default']['NAME']
            shutil.copy2(db_file, backup_file)
        
        logger.info(f"Database backup created: {backup_file}")
        return backup_file
        
    except Exception as e:
        logger.error(f"Failed to create database backup: {e}")
        return None
